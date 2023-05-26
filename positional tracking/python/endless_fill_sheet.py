########################################################################
#
# Copyright (c) 2022, STEREOLABS.
#
# All rights reserved.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS
# "AS IS" AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT
# LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR
# A PARTICULAR PURPOSE ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT
# OWNER OR CONTRIBUTORS BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL,
# SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT
# LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE,
# DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY
# THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT
# (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE
# OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.
#
########################################################################

import pyzed.sl as sl
import datetime
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from varname import nameof
import ImuData
import math
import keyboard

##
# Basic class to handle the timestamp of the different sensors to know if it is a new sensors_data or an old one
class TimestampHandler:
    def __init__(self):
        self.t_imu = sl.Timestamp()
        self.t_baro = sl.Timestamp()
        self.t_mag = sl.Timestamp()

    ##
    # check if the new timestamp is higher than the reference one, and if yes, save the current as reference
    def is_new(self, sensor):
        if isinstance(sensor, sl.IMUData):
            new_ = (sensor.timestamp.get_microseconds() > self.t_imu.get_microseconds())
            if new_:
                self.t_imu = sensor.timestamp
            return new_
        elif isinstance(sensor, sl.MagnetometerData):
            new_ = (sensor.timestamp.get_microseconds() > self.t_mag.get_microseconds())
            if new_:
                self.t_mag = sensor.timestamp
            return new_
        elif isinstance(sensor, sl.BarometerData):
            new_ = (sensor.timestamp.get_microseconds() > self.t_baro.get_microseconds())
            if new_:
                self.t_baro = sensor.timestamp
            return new_


def printSensorParameters(sensor_parameters):
    if sensor_parameters.is_available:
        print("*****************************")
        print("Sensor type: " + str(sensor_parameters.sensor_type))
        print("Max rate: "  + str(sensor_parameters.sampling_rate) + " "  + str(sl.SENSORS_UNIT.HERTZ))
        print("Range: "  + str(sensor_parameters.sensor_range) + " "  + str(sensor_parameters.sensor_unit))
        print("Resolution: " + str(sensor_parameters.resolution) + " "  + str(sensor_parameters.sensor_unit))
        if not math.isnan(sensor_parameters.noise_density):
            print("Noise Density: "  + str(sensor_parameters.noise_density) + " " + str(sensor_parameters.sensor_unit) + "/√Hz")
        if not math.isnan(sensor_parameters.random_walk):
            print("Random Walk: "  + str(sensor_parameters.random_walk) + " " + str(sensor_parameters.sensor_unit) + "/s/√Hz")

# def dot2comma(str):
def update_sheet():
    wb = load_workbook(filename="updateData.xlsx")
    ws = wb["grafici"]
    # fino a qui
    treeData = [["Type", "Leaf Color", "Height"], ["Maple", "Red", 549], ["Oak", "Green", 783], ["Pine", "Green", 1204]]
    for r in treeData:
        ws.append(r)
    wb.save("updateData.xlsx")

def write_xlsx(sheet_list):
    now = str(datetime.datetime.now().day) + '/'
    now += str(datetime.datetime.now().month) + '/'
    now += str(datetime.datetime.now().year) + ' '
    now += str(datetime.datetime.now().hour) + ':'
    mins = datetime.datetime.now().minute
    now += str(mins) if mins > 9 else '0' + str(mins)
    print(now)
    if len(sheet_list) == 0:
        return
    imu = sheet_list[0]

    wb = load_workbook(filename="updateData.xlsx")
    ws_dati = wb["dati"]
    wb.remove(ws_dati)
    ws_dati = wb.create_sheet("dati", 0)
    ws_dati.title = "dati"

    ws_dati.cell(column=1, row=1).value = 'v2'
    ws_dati.cell(column=2, row=1).value = now
    ws_dati.cell(column=1, row=2).value = nameof(imu.imu_Timestamp) + '[sec]'
    ws_dati.cell(column=2, row=2).value = nameof(imu.mag_Timestamp) + '[sec]'
    ws_dati.cell(column=3, row=2).value = nameof(imu.baro_Timestamp) + '[sec]'
    ws_dati.cell(column=4, row=2).value = nameof(imu.accX) + '[m/s^2]'
    ws_dati.cell(column=5, row=2).value = nameof(imu.accY) + '[m/s^2]'
    ws_dati.cell(column=6, row=2).value = nameof(imu.accZ) + '[m/s^2]'
    ws_dati.cell(column=7, row=2).value = nameof(imu.gyroX) + '[deg/s]'
    ws_dati.cell(column=8, row=2).value = nameof(imu.gyroY) + '[deg/S]'
    ws_dati.cell(column=9, row=2).value = nameof(imu.gyroZ) + '[deg/s]'
    ws_dati.cell(column=10, row=2).value = nameof(imu.magX) + '[uT]'
    ws_dati.cell(column=11, row=2).value = nameof(imu.magY) + '[uT]'
    ws_dati.cell(column=12, row=2).value = nameof(imu.magZ) + '[uT]'
    ws_dati.cell(column=13, row=2).value = nameof(imu.orX) + '[deg]'
    ws_dati.cell(column=14, row=2).value = nameof(imu.orY) + '[deg]'
    ws_dati.cell(column=15, row=2).value = nameof(imu.orZ) + '[deg]'
    ws_dati.cell(column=16, row=2).value = nameof(imu.press) + '[hPa]'
    ws_dati.cell(column=17, row=2).value = nameof(imu.rel_alt) + '[m]'
    ws_dati.cell(column=18, row=2).value = nameof(imu.moving)
    ws_dati.cell(column=19, row=2).value = nameof(imu.temp_left) + '[C]'
    ws_dati.cell(column=20, row=2).value = nameof(imu.temp_right) + '[C]'
    ws_dati.cell(column=21, row=2).value = nameof(imu.temp_imu) + '[C]'
    ws_dati.cell(column=22, row=2).value = nameof(imu.temp_barom) + '[C]'
    ws_dati['X2'] = 'delta accX'
    ws_dati['Y2'] = 'delta accY'
    ws_dati['Z2'] = 'delta accZ'

    offset_riga = 3
    riga = offset_riga
    for i in sheet_list:
        ws_dati['A'+str(riga)] = i.imu_Timestamp
        # ws_dati.cell(column=1, row=riga).value = i.imu_Timestamp
        ws_dati.cell(column=2, row=riga).value = i.mag_Timestamp
        ws_dati.cell(column=3, row=riga).value = i.baro_Timestamp
        ws_dati.cell(column=4, row=riga).value = i.accX
        ws_dati.cell(column=5, row=riga).value = i.accY
        ws_dati.cell(column=6, row=riga).value = i.accZ
        ws_dati.cell(column=7, row=riga).value = i.gyroX
        ws_dati.cell(column=8, row=riga).value = i.gyroY
        ws_dati.cell(column=9, row=riga).value = i.gyroZ
        ws_dati.cell(column=10, row=riga).value = i.magX
        ws_dati.cell(column=11, row=riga).value = i.magY
        ws_dati.cell(column=12, row=riga).value = i.magZ
        ws_dati.cell(column=13, row=riga).value = i.orX
        ws_dati.cell(column=14, row=riga).value = i.orY
        ws_dati.cell(column=15, row=riga).value = i.orZ
        ws_dati.cell(column=16, row=riga).value = i.press
        ws_dati.cell(column=17, row=riga).value = i.rel_alt
        ws_dati.cell(column=18, row=riga).value = i.moving
        ws_dati.cell(column=19, row=riga).value = i.temp_left
        ws_dati.cell(column=20, row=riga).value = i.temp_right
        ws_dati.cell(column=21, row=riga).value = i.temp_imu
        ws_dati.cell(column=22, row=riga).value = i.temp_barom

        if not riga == offset_riga:
            end = str(riga) + '*(A' + str(riga) + '-A' + str(riga-1) + ')^2'
            ws_dati['X'+str(riga)] = '=0.5*D' + end
            ws_dati['Y'+str(riga)] = '=0.5*E' + end
            ws_dati['Z'+str(riga)] = '=0.5*F' + end
        riga += 1

    wb.save("updateData.xlsx")


def main():
    # Create a Camera object
    zed = sl.Camera()
    init_params = sl.InitParameters()
    init_params.depth_mode = sl.DEPTH_MODE.NONE

    # Open the camera
    err = zed.open(init_params)
    if err != sl.ERROR_CODE.SUCCESS:
        print(repr(err))
        zed.close()
        exit(1)

    # Get camera information sensors_data
    info = zed.get_camera_information()
    cam_model = info.camera_model
    if cam_model == sl.MODEL.ZED:
        print("This tutorial only supports ZED-M and ZED2 camera models, ZED does not have additional sensors")
        exit(1)

    """
    printSensorParameters(info.sensors_configuration.accelerometer_parameters)  # accelerometer configuration
    printSensorParameters(info.sensors_configuration.gyroscope_parameters)  # gyroscope configuration
    printSensorParameters(info.sensors_configuration.magnetometer_parameters)  # magnetometer configuration
    printSensorParameters(info.sensors_configuration.barometer_parameters)  # barometer configuration
    """
    # Used to store the sensors timestamp to know if the sensors_data is a new one or not
    ts_handler = TimestampHandler()

    data_list = []
    sensors_data = sl.SensorsData()
    i = 0
    print("go!!")
    while True:
        if keyboard.is_pressed('enter'):
            break
        if zed.get_sensors_data(sensors_data, sl.TIME_REFERENCE.CURRENT) == sl.ERROR_CODE.SUCCESS:
            # Check if the data has been updated since the last time. IMU is the sensor with the highest rate
            imu_data = sensors_data.get_imu_data()
            if ts_handler.is_new(imu_data):
                if i % 500 == 0:
                    print(i)
                # da prendere: quaternioni, translation e imu_orientation
                #print(imu_data.timestamp.data_ns)
                #imu_or = imu_data.get_pose().get_orientation().get()
                imu_or = imu_data.get_pose().get_euler_angles(radian=False)
                imu_vel = imu_data.get_angular_velocity()
                imu_acc = imu_data.get_linear_acceleration()
                mag_data = sensors_data.get_magnetometer_data().get_magnetic_field_calibrated()
                # non restituisce timestamp i dati nella forma giusta
                #imuT = 0.000000001 * imu_data.timestamp.data_ns
                imuT = imu_data.timestamp.get_milliseconds()
                # print(imuT)
                #magT = 0.000000001 * sensors_data.get_magnetometer_data().timestamp.data_ns
                magT = sensors_data.get_magnetometer_data().timestamp.get_milliseconds()
                #barT = 0.000000001 * sensors_data.get_barometer_data().timestamp.data_ns
                barT = sensors_data.get_barometer_data().timestamp.get_milliseconds()
                accX = imu_acc[0]
                accY = imu_acc[1]
                accZ = imu_acc[2]
                gyrX = imu_vel[0]
                gyrY = imu_vel[1]
                gyrZ = imu_vel[2]
                magX = mag_data[0]
                magY = mag_data[1]
                magZ = mag_data[2]
                orX = imu_or[0]
                orY = imu_or[1]
                orZ = imu_or[2]
                #prova
                press = sensors_data.get_barometer_data().pressure
                # ? sembra che altitudine relativa sia sempre 0.0, nonostante sposti la telecamera in alto
                r_alt = sensors_data.get_barometer_data().relative_altitude
                # non restituisce il formato giusto. Inoltre quando sta ferma sembra che dica che si muove
                # mov è 0 se telecamera è ferma, 1 se si
                #
                # muove e -1 se sta cadendo
                mov = 0 if sensors_data.camera_moving_state == sl.CAMERA_MOTION_STATE.STATIC else 1 if sensors_data.camera_moving_state == sl.CAMERA_MOTION_STATE.MOVING else -1
                tLeft = sensors_data.get_temperature_data().get(sl.SENSOR_LOCATION.ONBOARD_LEFT)
                tRight = sensors_data.get_temperature_data().get(sl.SENSOR_LOCATION.ONBOARD_RIGHT)
                tImu = sensors_data.get_temperature_data().get(sl.SENSOR_LOCATION.IMU)
                tBar = sensors_data.get_temperature_data().get(sl.SENSOR_LOCATION.BAROMETER)
                data_list.append(ImuData.ImuData(imuT, magT, barT, accX, accY, accZ, gyrX, gyrY, gyrZ, magX, magY, magZ, orX, orY, orZ, press, r_alt, mov, tLeft, tRight, tImu, tBar))
                i += 1
    print(len(data_list))
    write_xlsx(data_list)
    zed.close()
    return 0


if __name__ == "__main__":
    main()

# =SCARTO(sides!$E$3;0;0;CONTA.VALORI(dati!$E:$E)-1;1)